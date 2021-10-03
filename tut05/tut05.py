import csv
import os
from openpyxl import Workbook, load_workbook

#Helper Function to insert Headings in Overall Marksheet
def insert_headings (sheet,rno,name):
    
    sheet['A1'] = "Roll No."
    sheet['B1'] = rno
    sheet['A2'] = "Name of Student"
    sheet['B2'] = name
    sheet['A3'] = "Discipline"
    sheet['B3'] = rno[4:6]
    sheet['A4'] = "Semester No."
    sheet['A5'] = "Semester wise Credit Taken"
    sheet['A6'] = "SPI"
    sheet['A7'] = "Total Credits taken"
    sheet['A8'] = "CPI"
    return

#Function to create .xlsx file of given rno
#Using Relevant Dictionaries
def generate_marksheet(rno, marksheet, course_map, roll_map, grade_map):

    op_file = "output\\"+rno+".xlsx"

    wb = Workbook()

    #Create and initialize Overall marksheet
    overall_sheet = wb.active
    overall_sheet.title = "Overall"
    insert_headings(overall_sheet,rno,roll_map[rno])

    total_credits = 0
    total_grade = 0
    cpi = 0

    #Create semester marksheets
    for sem in marksheet[rno]:

        sheet_name = "Sem" + str(sem)
        wb.create_sheet(sheet_name)

        sem_credits = 0
        sem_grade = 0
        spi = 0

        curr_sheet = wb[sheet_name]

        heading = ["Sr. No","Course No.","Course Name","L-T-P","Credits","Course Type","Grade"]

        curr_sheet.append(heading)
        srno = 1

        #Append new row for every course in that sem
        for course in marksheet[rno][sem]:
            new_row = [srno, course]
            new_row.extend(course_map[course])
            new_row.extend(marksheet[rno][sem][course])

            curr_sheet.append(new_row)
            srno += 1
         
            #Keep track of spi
            sem_credits += int(new_row[4])
            sem_grade += grade_map[new_row[6]] * int(new_row[4])

        spi = sem_grade/sem_credits

        #Update cpi
        total_credits += sem_credits
        total_grade += sem_grade
        cpi = total_grade/total_credits

        Sem = int(sem)+1

        #Insert Data of the current sem in overall Marksheet 
        overall_sheet.cell(row = 4, column = Sem).value = sem
        overall_sheet.cell(row = 5, column = Sem).value = sem_credits
        overall_sheet.cell(row = 6, column = Sem).value = spi
        overall_sheet.cell(row = 7, column = Sem).value = total_credits
        overall_sheet.cell(row = 8, column = Sem).value = cpi


    wb.save(op_file)

    return

#{Course Code : Course Name}
course_map = {}

#{Roll Number : Name of Student}
roll_map = {}

#{Roll Number : {Semester No. : {Course : Type, Grade}}}
# marksheet[roll][semester][course] = [Course Type, Course Grade]
marksheet = {}

#Build Course Dict
with open("subjects_master.csv","r") as sub_file:
    reader = csv.reader(sub_file)
    heading = next(reader)

    for row in reader:
        course_map[row[0]] = [row[1],row[2],row[3]]

#Build Roll number dict
with open("names-roll.csv","r") as stud_file:
    reader = csv.reader(stud_file)
    heading = next(reader)

    for row in reader:
        roll_map[row[0]] = row[1] 

#Build Marksheet (As Dict)
with open("grades.csv","r") as grades:
    reader = csv.reader(grades)
    heading = next(reader)

    for row in reader:

        rno = row[0]
        sem = row[1]
        sub = row[2]
        grade = row[4]
        sub_type = row[5]
        grade = grade.strip()

        if grade[-1] == '*':
            grade = grade[:-1]

        if rno not in marksheet:
            marksheet[rno] = {}

        if sem not in marksheet[rno]:
            marksheet[rno][sem] = {}

        marksheet[rno][sem][sub] = [sub_type,grade]

if not os.path.exists("output"):
    os.mkdir("output")

grade_map = {
                "AA" : 10, 
                "AB" : 9,
                "BB" : 8,
                "BC" : 7,
                "CC" : 6,
                "CD" : 5,
                "DD" : 4,
                "F"  : 0,
                "I"  : 0
            }

for rno in marksheet:
    generate_marksheet(rno,marksheet,course_map,roll_map,grade_map)