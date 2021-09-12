import csv
import os
from openpyxl import Workbook, load_workbook

#Function to keep only relevant columns from a row
def clean(raw_row, relevant_cols):
    cleaned_row = []
    for col in relevant_cols:
        cleaned_row.append(raw_row[col])
    return cleaned_row

#Function takes in a row and puts it in proper file
def output(row,heading,output_folder,output_field):
    #Detect faulty rows
    if len(row[output_field]) == 0: return

    #Get relative path of output file
    output_file = output_folder + row[output_field] + ".xlsx"
    
    #If file doesn't exist, create and put headings
    if os.path.isfile(output_file) == False:
        workb = Workbook()
        curr_sheet = workb.active
        curr_sheet.append(heading)
        workb.save(output_file)

    #Append the row to the file
    workb = load_workbook(filename = output_file)
    curr_sheet = workb.active
    curr_sheet.append(row)
    workb.save(output_file)
    return

#Utility lists and strings
output_roll = "output_individual_roll\\"
output_sub = "output_by_subject\\"
relevant_cols = [0,1,3,8];

with open('regtable_old.csv', 'r') as data_file:
    reader = csv.reader(data_file)

    #Read first line and extract relevant headings
    heading = next(reader)
    heading = clean(heading,relevant_cols)

    #Append every row in proper files
    for row in reader:
        if len(row)>8:   #Check if row is of proper size
            row = clean(row,relevant_cols)
            output(row,heading,output_roll,0)
            output(row,heading,output_sub,2)
